from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from datetime import date
import io, openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from database import get_db
from auth import get_current_user
import models, schemas

router = APIRouter(prefix="/payments", tags=["Pagos"])


# ─── Listar (con filtros) ─────────────────
@router.get("/", response_model=List[schemas.PaymentOut])
def list_payments(
    category: Optional[str] = Query(None),
    status:   Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to:   Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    q = db.query(models.Payment).filter(models.Payment.user_id == current_user.id)
    if category:
        q = q.filter(models.Payment.category == category)
    if status:
        q = q.filter(models.Payment.status == status)
    if date_from:
        q = q.filter(models.Payment.date >= date_from)
    if date_to:
        q = q.filter(models.Payment.date <= date_to)
    return q.order_by(models.Payment.date.desc()).all()


# ─── Crear ────────────────────────────────
@router.post("/", response_model=schemas.PaymentOut, status_code=201)
def create_payment(
    data: schemas.PaymentCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    payment = models.Payment(**data.model_dump(), user_id=current_user.id)
    db.add(payment)
    db.commit()
    db.refresh(payment)
    return payment


# ─── Obtener uno ──────────────────────────
@router.get("/{payment_id}", response_model=schemas.PaymentOut)
def get_payment(
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    payment = db.query(models.Payment).filter(
        models.Payment.id == payment_id,
        models.Payment.user_id == current_user.id
    ).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    return payment


# ─── Actualizar ───────────────────────────
@router.put("/{payment_id}", response_model=schemas.PaymentOut)
def update_payment(
    payment_id: int,
    data: schemas.PaymentUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    payment = db.query(models.Payment).filter(
        models.Payment.id == payment_id,
        models.Payment.user_id == current_user.id
    ).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    for field, value in data.model_dump().items():
        setattr(payment, field, value)
    db.commit()
    db.refresh(payment)
    return payment


# ─── Eliminar ─────────────────────────────
@router.delete("/{payment_id}", status_code=204)
def delete_payment(
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    payment = db.query(models.Payment).filter(
        models.Payment.id == payment_id,
        models.Payment.user_id == current_user.id
    ).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    db.delete(payment)
    db.commit()


# ─── Estadísticas para gráficas ───────────
@router.get("/stats/summary")
def get_stats(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    payments = db.query(models.Payment).filter(
        models.Payment.user_id == current_user.id
    ).all()

    total = sum(p.amount for p in payments)
    total_pagado = sum(p.amount for p in payments if p.status == "Pagado")
    total_pendiente = sum(p.amount for p in payments if p.status == "Pendiente")

    # por categoría
    by_category = {}
    for p in payments:
        by_category[p.category] = by_category.get(p.category, 0) + p.amount

    # por mes
    by_month = {}
    for p in payments:
        key = p.date.strftime("%Y-%m")
        by_month[key] = by_month.get(key, 0) + p.amount

    return {
        "total": round(total, 2),
        "total_pagado": round(total_pagado, 2),
        "total_pendiente": round(total_pendiente, 2),
        "count": len(payments),
        "by_category": [{"category": k, "amount": round(v, 2)} for k, v in by_category.items()],
        "by_month": [{"month": k, "amount": round(v, 2)} for k, v in sorted(by_month.items())],
    }


# ─── Exportar Excel ───────────────────────
@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    payments = db.query(models.Payment).filter(
        models.Payment.user_id == current_user.id
    ).order_by(models.Payment.date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"

    headers = ["ID", "Descripción", "Monto (₡)", "Categoría", "Estado", "Fecha"]
    ws.append(headers)

    for p in payments:
        ws.append([p.id, p.description, p.amount, p.category, p.status, str(p.date)])

    # estilo encabezado
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill("solid", fgColor="7C6AF7")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    # ancho automático
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pagos.xlsx"}
    )


# ─── Exportar PDF ─────────────────────────
@router.get("/export/pdf")
def export_pdf(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    payments = db.query(models.Payment).filter(
        models.Payment.user_id == current_user.id
    ).order_by(models.Payment.date.desc()).all()

    stream = io.BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Reporte de Pagos — {current_user.username}", styles["Title"]))
    elements.append(Spacer(1, 12))

    data = [["Descripción", "Monto", "Categoría", "Estado", "Fecha"]]
    for p in payments:
        data.append([p.description, f"₡{p.amount:,.2f}", p.category, p.status, str(p.date)])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7C6AF7")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F0FF")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(table)

    total = sum(p.amount for p in payments)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"<b>Total: ₡{total:,.2f}</b>", styles["Normal"]))

    doc.build(elements)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=pagos.pdf"}
    )
